import openpyxl

class Excel_utils :
    @staticmethod
    def read_data_from_excel(file_path, sheet_name, row_number, column_number):
        excel_file = openpyxl.load_workbook(file_path)
        sheet = excel_file[sheet_name]
        return sheet.cell(row=row_number, column=column_number).value

    @staticmethod
    def write_data_to_excel(file_path, sheet_name, row_number, column_number, data):
        excel_file = openpyxl.load_workbook(file_path)
        sheet = excel_file[sheet_name]
        sheet.cell(row=row_number, column=column_number).value = data
        excel_file.save(file_path)

    @staticmethod
    def  get_row_count(file_path, sheet_name):
        excel_file = openpyxl.load_workbook(file_path)
        sheet = excel_file[sheet_name]
        return sheet.max_row